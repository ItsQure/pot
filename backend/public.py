# Author: Ankush Patel
# Description: Converts Excel file to JSON file and pushes it
# to Firebase server.

import config
import datetime
import pyrebase
import openpyxl
import json

cfg = {
    "apiKey": "AIzaSyBcSBszeaGa0i3lxoLTd52sMOPXeG3a0Pc",
    "authDomain": "pothole-233315.firebaseapp.com",
    "databaseURL": "https://pothole-233315.firebaseio.com",
    "storageBucket": "pothole-233315.appspot.com",
    "serviceAccount": config.API_PATH
}

workOrder = {
    "woDate": "",
    "woNum": "",
    "location": "",
    "priority": "",
    "description": ""
}

# Requires service account credentials
firebase = pyrebase.initialize_app(cfg)
db = firebase.database()

wb = openpyxl.load_workbook('open_potholes.xlsx')
ws = wb['Sheet1'] # There is only one sheet

# Data headings are on row 3
woDate = ws['A3'].column - 1
woNum = ws['E3'].column - 1
location = ws['C3'].column - 1
priority = 3 # Priority for public reported potholes are maximum
description = ''

# Grabs data in the first row
for row in ws.iter_rows(min_row=4, min_col=1, max_col=8):
    time = row[woDate].value # Used as key in database
    workOrder["woDate"] = time.strftime('%m/%d/%Y')
    workOrder["woNum"] = row[woNum].value
    workOrder["location"] = row[location].value

    # Converts file to JSON
    json_data = json.dumps(workOrder)
    json_data = json.loads(json_data)
    db.child("workOrders").push(json_data)

wos = db.child("workOrders").get().val()
json_data = json.dumps(wos)
json_data = json.loads(json_data)

times = list(json_data.keys())

for i in times:
    print(json_data[i])
